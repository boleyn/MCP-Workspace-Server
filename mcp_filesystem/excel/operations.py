"""Excel operations for the MCP filesystem server."""

import json
import os
import shutil
import time
from datetime import datetime
from functools import partial
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import anyio
from mcp.server.fastmcp.utilities.logging import get_logger
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series, ScatterChart
from openpyxl.formula import Tokenizer
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
    range_boundaries,
)

from ..security import PathValidator
from ..operations import format_timestamp
from .converters import read_csv, write_csv
from .utils import format_read_result, format_size, normalize_output_format

logger = get_logger(__name__)

# Force INFO level logging for debugging
import logging
logger.setLevel(logging.INFO)


class ExcelOperations:
    """Excel file operations with path validation."""

    def __init__(
        self,
        validator: PathValidator,
        config: Optional[Dict[str, Any]] = None,
    ):
        self.validator = validator
        config = config or {}
        self.default_max_rows: int = int(
            os.environ.get("MCP_EXCEL_MAX_ROWS", config.get("default_max_rows", 1000))
        )
        self.max_file_size_mb: float = float(
            os.environ.get("MCP_EXCEL_MAX_SIZE_MB", config.get("max_file_size_mb", 50))
        )
        self.supported_formats: List[str] = [
            fmt.lower() for fmt in config.get("supported_formats", [".xlsx", ".xls", ".csv"])
        ]
        self.enable_formulas: bool = bool(config.get("enable_formulas", True))
        self._project_root = Path(__file__).resolve().parent.parent.parent
        self.templates_config_path: Optional[Path] = None
        self.templates: List[Dict[str, Any]] = []
        self._template_error: Optional[str] = None
        self._load_templates_config(config)

    async def read_excel(
        self,
        path: str,
        sheet: Optional[str] = None,
        range_str: Optional[str] = None,
        max_rows: Optional[int] = None,
        output_format: str = "markdown",
    ) -> Dict[str, Any]:
        """Read an Excel file and return structured data.

        Args:
            path: Virtual path string to the Excel file (NOT a Path object)
            sheet: Sheet name (None or empty string means use first sheet)
            range_str: Cell range (e.g., "A1:C10") or None for entire sheet
        """
        # Normalize sheet parameter: empty string -> None
        if sheet == "":
            sheet = None
        
        stage_start = time.perf_counter()

        def _log_stage(label: str) -> None:
            nonlocal stage_start
            now = time.perf_counter()
            logger.info(f"read_excel: {label} took {now - stage_start:.3f}s")
            stage_start = now

        logger.info(f"read_excel: Starting to read Excel file: path={path}, sheet={sheet}, range={range_str}, max_rows={max_rows}, output_format={output_format}")
        
        if not isinstance(path, str):
            error_msg = f"read_excel expects str (virtual path), got {type(path).__name__}"
            logger.error(f"read_excel: {error_msg}")
            raise TypeError(error_msg)

        try:
            logger.info(f"read_excel: Validating Excel path: {path}")
            abs_path = await self._validated_excel_path(path)
            logger.info(f"read_excel: Validated path: {abs_path}")
            _log_stage("path validation")
        except Exception as e:
            logger.error(f"read_excel: Path validation failed for '{path}': {type(e).__name__}: {e}", exc_info=True)
            raise
        
        try:
            logger.info(f"read_excel: Checking file size: {abs_path}")
            self._check_file_size(abs_path)
            logger.info(f"read_excel: File size check passed")
            _log_stage("size check")
        except Exception as e:
            logger.error(f"read_excel: File size check failed for '{path}': {type(e).__name__}: {e}", exc_info=True)
            raise

        if abs_path.suffix.lower() == ".csv":
            return await self._read_csv_file(abs_path, sheet, range_str, max_rows, output_format)

        # 先读取值（data_only=True），获取计算结果
        wb_values = None
        ws_values = None
        try:
            logger.info(f"Loading workbook with data_only=True: {abs_path}")
            wb_values = await self._load_workbook(abs_path, read_only=True, data_only=True)
            logger.info(f"Workbook loaded, available sheets: {wb_values.sheetnames}")
            ws_values = self._get_sheet(wb_values, sheet)
            logger.info(f"Successfully loaded workbook with data_only=True, sheet: {ws_values.title}")
            _log_stage("load workbook data_only=True")
        except Exception as e:
            logger.warning(f"Failed to load workbook with data_only=True: {type(e).__name__}: {e}", exc_info=True)
            # 如果失败，回退到 data_only=False
            try:
                logger.info(f"Falling back to data_only=False: {abs_path}")
                wb_values = await self._load_workbook(abs_path, read_only=True, data_only=False)
                logger.info(f"Workbook loaded (data_only=False), available sheets: {wb_values.sheetnames}")
                ws_values = self._get_sheet(wb_values, sheet)
                logger.info(f"Successfully loaded workbook with data_only=False, sheet: {ws_values.title}")
                _log_stage("load workbook data_only=False (fallback)")
            except Exception as e2:
                logger.error(f"Failed to load workbook even with data_only=False: {type(e2).__name__}: {e2}", exc_info=True)
                raise

        try:
            min_row, max_row, min_col, max_col = self._parse_range(ws_values, range_str)
            total_rows = max(0, max_row - min_row + 1)
            logger.debug(f"Parsed range: min_row={min_row}, max_row={max_row}, min_col={min_col}, max_col={max_col}, total_rows={total_rows}")
            _log_stage("parse range")

            row_limit = max_rows or self.default_max_rows
            row_limit = row_limit if row_limit > 0 else self.default_max_rows

            rows: List[List[Any]] = []
            truncated = False
            iter_start = time.perf_counter()

            # 读取公式（data_only=False）仅在行数不大于 100 且启用了公式模式时进行
            wb_formulas = None
            ws_formulas = None
            use_separate_formulas = False
            formula_enabled = self.enable_formulas and total_rows <= 100

            if formula_enabled:
                try:
                    logger.info(
                        "Attempting to load workbook with data_only=False for formulas: %s (total_rows=%d)",
                        abs_path,
                        total_rows,
                    )
                    wb_formulas = await self._load_workbook(abs_path, read_only=True, data_only=False)
                    ws_formulas = self._get_sheet(wb_formulas, sheet)
                    use_separate_formulas = True
                    logger.info(
                        "Successfully loaded separate workbook for formulas, sheet: %s",
                        ws_formulas.title,
                    )
                    _log_stage("load workbook for formulas (data_only=False)")
                except Exception as e:
                    logger.warning(
                        "Failed to load separate workbook for formulas (will use values workbook): %s: %s",
                        type(e).__name__,
                        e,
                        exc_info=True,
                    )
                    wb_formulas = wb_values
                    ws_formulas = ws_values
                    use_separate_formulas = False
                    logger.info("Using values workbook for formulas as fallback")
            else:
                wb_formulas = wb_values
                ws_formulas = ws_values
                use_separate_formulas = False
                reason = "enable_formulas=False" if not self.enable_formulas else f"row_count_exceeds_limit({total_rows}>100)"
                logger.info("Formula workbook loading skipped (%s)", reason)

            values_iter = ws_values.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )

            if use_separate_formulas:
                formula_iter = ws_formulas.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=False,
                )
                for row_idx, (val_row, formula_row) in enumerate(zip(values_iter, formula_iter), start=1):
                    if row_idx > row_limit:
                        truncated = True
                        break
                    row_data: List[Any] = []
                    for val_cell, formula_cell in zip(val_row, formula_row):
                        if val_cell is None and getattr(formula_cell, "data_type", "") == "f" and formula_cell.value:
                            row_data.append(formula_cell.value)
                        else:
                            row_data.append(val_cell)
                    rows.append(row_data)
            else:
                for row_idx, val_row in enumerate(values_iter, start=1):
                    if row_idx > row_limit:
                        truncated = True
                        break
                    rows.append(list(val_row))

            result: Dict[str, Any] = {
                "sheet": ws_values.title,
                "sheets": wb_values.sheetnames,  # 添加所有工作表列表
                "rows": rows,
                "total_rows": total_rows,
                "returned_rows": len(rows),
                "truncated": truncated,
            }
            logger.info(
                "read_excel: iterated rows %d/%d (truncated=%s) in %.3fs",
                len(rows),
                total_rows,
                truncated,
                time.perf_counter() - iter_start,
            )
            return result
        finally:
            # 确保关闭 workbook
            try:
                if wb_values is not None:
                    wb_values.close()
                    logger.info("Closed values workbook")
            except Exception as e:
                logger.warning(f"Error closing values workbook: {e}")
            try:
                if use_separate_formulas and wb_formulas is not None and wb_formulas != wb_values:
                    wb_formulas.close()
                    logger.info("Closed formulas workbook")
            except Exception as e:
                logger.warning(f"Error closing formulas workbook: {e}")
            # 如果没有使用独立的 formulas workbook，只需要关闭一个
            if not use_separate_formulas and wb_formulas is not None and wb_formulas != wb_values:
                try:
                    wb_formulas.close()
                    logger.info("Closed workbook")
                except Exception as e:
                    logger.warning(f"Error closing workbook: {e}")


    async def write_excel(
        self,
        path: str,
        data: List[List[Any]],
        sheet: str = "Sheet1",
        overwrite: bool = False,
    ) -> Dict[str, Any]:
        """Create or overwrite an Excel file.

        Args:
            path: Virtual path string to the Excel file (NOT a Path object)
            data: 2D array of data. If you want headers, include them as the first row.
        """
        if not isinstance(path, str):
            raise TypeError(f"write_excel expects str (virtual path), got {type(path).__name__}")

        abs_path = await self._validated_excel_path(path)

        if abs_path.exists() and not overwrite:
            raise FileExistsError(f"File already exists: {path}")

        if not abs_path.parent.exists():
            await anyio.to_thread.run_sync(lambda: abs_path.parent.mkdir(parents=True, exist_ok=True))

        wb = Workbook()
        ws = wb.active
        ws.title = sheet

        for row in data:
            ws.append(row)

        await anyio.to_thread.run_sync(wb.save, abs_path)

        return {
            "success": True,
            "path": self.validator.real_to_virtual(abs_path),
            "sheet": sheet,
            "rows_written": len(data),
            "message": "Excel file created successfully",
        }

    async def update_excel_cells(
        self,
        path: str,
        updates: List[Dict[str, Any]],
        sheet: Optional[str] = None,
        create_if_missing: bool = False,
    ) -> Dict[str, Any]:
        """Update specific cells or ranges in an Excel file."""
        abs_path, allowed = await self.validator.validate_path(path)
        if not allowed:
            raise ValueError(f"Path outside allowed directories: {path}")

        if not abs_path.exists():
            if not create_if_missing:
                raise FileNotFoundError(f"File not found: {path}")
            wb = Workbook()
        else:
            wb = await self._load_workbook(abs_path, read_only=False, data_only=False)

        if sheet:
            ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(title=sheet)
        else:
            ws = wb.active

        updated_cells = 0
        for update in updates:
            updated_cells += self._apply_update(ws, update)

        await anyio.to_thread.run_sync(wb.save, abs_path)

        return {
            "success": True,
            "path": self.validator.real_to_virtual(abs_path),
            "updated_cells": updated_cells,
            "message": "Cells updated successfully",
        }

    async def list_sheets(self, path: str) -> Dict[str, Any]:
        """List worksheets with basic statistics."""
        abs_path = await self._validated_excel_path(path, must_exist=True)
        if self._is_csv(abs_path):
            rows = await read_csv(abs_path)
            rows_count = len(rows)
            cols_count = max((len(r) for r in rows), default=0)
            sheet_info = {
                "name": abs_path.stem or "Sheet1",
                "index": 0,
                "rows": rows_count,
                "columns": cols_count,
                "is_active": True,
            }
            return {
                "path": self.validator.real_to_virtual(abs_path),
                "sheets": [sheet_info],
                "total_sheets": 1,
            }

        wb = await self._load_workbook(abs_path, read_only=True, data_only=True)

        sheets: List[Dict[str, Any]] = []
        active_title = wb.active.title
        for idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            sheets.append(
                {
                    "name": sheet_name,
                    "index": idx,
                    "rows": ws.max_row,
                    "columns": ws.max_column,
                    "is_active": sheet_name == active_title,
                }
            )

        return {
            "path": self.validator.real_to_virtual(abs_path),
            "sheets": sheets,
            "total_sheets": len(sheets),
        }

    async def get_excel_info(self, path: str) -> Dict[str, Any]:
        """Get metadata for an Excel file."""
        abs_path = await self._validated_excel_path(path, must_exist=True)
        self._check_file_size(abs_path)

        stat = abs_path.stat()

        if self._is_csv(abs_path):
            rows = await read_csv(abs_path)
            rows_count = len(rows)
            cols_count = max((len(r) for r in rows), default=0)
            return {
                "path": self.validator.real_to_virtual(abs_path),
                "file_size": format_size(stat.st_size),
                "created_time": format_timestamp(stat.st_ctime),
                "modified_time": format_timestamp(stat.st_mtime),
                "sheets": [
                    {"name": abs_path.stem or "Sheet1", "rows": rows_count, "columns": cols_count}
                ],
                "total_sheets": 1,
                "has_formulas": False,
            }

        wb = await self._load_workbook(abs_path, read_only=True, data_only=False)

        sheets_info = []
        has_formulas = False
        formula_scan_limit = 2000
        scanned_cells = 0

        for ws in wb.worksheets:
            sheets_info.append(
                {
                    "name": ws.title,
                    "rows": ws.max_row,
                    "columns": ws.max_column,
                }
            )
            if not has_formulas and self.enable_formulas:
                for row in ws.iter_rows(values_only=False):
                    for cell in row:
                        scanned_cells += 1
                        if cell.data_type == "f":
                            has_formulas = True
                            break
                    if has_formulas or scanned_cells >= formula_scan_limit:
                        break
            if has_formulas or scanned_cells >= formula_scan_limit:
                break

        return {
            "path": self.validator.real_to_virtual(abs_path),
            "file_size": format_size(stat.st_size),
            "created_time": format_timestamp(stat.st_ctime),
            "modified_time": format_timestamp(stat.st_mtime),
            "sheets": sheets_info,
            "total_sheets": len(sheets_info),
            "has_formulas": has_formulas,
        }

    async def excel_to_csv(
        self,
        path: str,
        sheet: Optional[str] = None,
        output_path: Optional[str] = None,
        delimiter: str = ",",
        encoding: str = "utf-8",
    ) -> Dict[str, Any]:
        """Convert an Excel sheet to CSV."""
        abs_path = await self._validated_excel_path(path, must_exist=True)
        self._check_file_size(abs_path)
        if self._is_csv(abs_path):
            raise ValueError("Input file is already a CSV file")
        wb = await self._load_workbook(abs_path, read_only=True, data_only=True)
        ws = self._get_sheet(wb, sheet)

        target_path = await self._resolve_output_path(
            source_path=abs_path, provided_output=output_path, default_suffix=".csv"
        )

        rows = ws.iter_rows(values_only=True)
        await write_csv(target_path, rows, delimiter=delimiter, encoding=encoding)

        return {
            "success": True,
            "path": self.validator.real_to_virtual(abs_path),
            "output_path": self.validator.real_to_virtual(target_path),
            "sheet": ws.title,
            "message": "Excel exported to CSV successfully",
        }

    async def csv_to_excel(
        self,
        path: str,
        output_path: Optional[str] = None,
        sheet: str = "Sheet1",
        delimiter: str = ",",
        encoding: str = "utf-8",
        overwrite: bool = False,
    ) -> Dict[str, Any]:
        """Convert a CSV file to Excel."""
        abs_path, allowed = await self.validator.validate_path(path)
        if not allowed:
            raise ValueError(f"Path outside allowed directories: {path}")
        if not abs_path.exists():
            raise FileNotFoundError(f"CSV file not found: {path}")
        if not self._is_csv(abs_path):
            raise ValueError("csv_to_excel expects a CSV input file")

        target_path = await self._resolve_output_path(
            source_path=abs_path, provided_output=output_path, default_suffix=".xlsx"
        )

        if target_path.exists() and not overwrite:
            raise FileExistsError(f"File already exists: {target_path}")

        rows = await read_csv(abs_path, delimiter=delimiter, encoding=encoding)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet
        for row in rows:
            ws.append(row)

        await anyio.to_thread.run_sync(wb.save, target_path)

        return {
            "success": True,
            "path": self.validator.real_to_virtual(abs_path),
            "output_path": self.validator.real_to_virtual(target_path),
            "sheet": sheet,
            "rows_written": len(rows),
            "message": "CSV converted to Excel successfully",
        }

    async def list_templates(self) -> Dict[str, Any]:
        """List configured Excel templates (title + description only)."""
        self._ensure_templates_available()
        templates = [{"title": tpl["title"], "desc": tpl.get("desc", "")} for tpl in self.templates]
        return {"success": True, "templates": templates, "count": len(templates)}

    async def create_from_template(
        self,
        template_title: str,
        directory: Optional[str] = None,
        file_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Create a new Excel file from a configured template."""
        self._ensure_templates_available()
        template = self._get_template_by_title(template_title)
        source_path: Path = template["source"]

        if not source_path.exists():
            raise FileNotFoundError(f"Template source not found: {self._friendly_path(source_path)}")

        self._ensure_supported_extension(source_path)

        target_path = await self._resolve_template_target_path(
            directory=directory,
            file_name=file_name,
            template_source=source_path,
            template_title=template["title"],
        )
        unique_target = await self._make_unique_path(target_path)

        if not unique_target.parent.exists():
            await anyio.to_thread.run_sync(lambda: unique_target.parent.mkdir(parents=True, exist_ok=True))

        await anyio.to_thread.run_sync(shutil.copy2, source_path, unique_target)

        # Convert to virtual path
        virtual_path = self.validator.real_to_virtual(unique_target)

        # Debug logging
        logger.debug(f"create_from_template: real_path={unique_target}, virtual_root={self.validator.virtual_root}, virtual_path={virtual_path}")

        return {
            "success": True,
            "path": virtual_path,
            "file_name": unique_target.name,
            "message": "Excel file created from template",
        }

    # ---------- Helpers ----------

    def _load_templates_config(self, config: Dict[str, Any]) -> None:
        """Load template definitions from config or templates.json."""
        templates_data = config.get("templates")
        templates_file = config.get("templates_file")

        templates_path: Optional[Path] = None
        if templates_file:
            try:
                templates_path = self._resolve_config_path(templates_file, self._project_root)
            except Exception as e:
                self._template_error = f"Invalid templates_file path: {e}"
                logger.warning(self._template_error)
                return
        else:
            templates_path = self._project_root / "excel_templates" / "templates.json"

        self.templates_config_path = templates_path

        if templates_data is None:
            if templates_path and templates_path.exists():
                try:
                    with open(templates_path, "r", encoding="utf-8") as f:
                        templates_data = json.load(f)
                except Exception as e:
                    self._template_error = f"Failed to load Excel templates: {e}"
                    logger.warning(self._template_error)
                    return
            else:
                self.templates = []
                self._template_error = (
                    f"No Excel templates configured (expected {self._friendly_path(templates_path)})"
                )
                return

        try:
            self.templates = self._normalize_templates(templates_data, templates_path)
            self._template_error = None
            if self.templates:
                logger.info(
                    "Loaded %d Excel templates from %s",
                    len(self.templates),
                    self._friendly_path(templates_path) if templates_path else "config",
                )
        except Exception as e:
            self._template_error = str(e)
            self.templates = []

    def _normalize_templates(
        self, templates_data: Any, config_path: Optional[Path]
    ) -> List[Dict[str, Any]]:
        """Normalize template entries and resolve paths."""
        if templates_data is None:
            return []
        if not isinstance(templates_data, list):
            raise ValueError("Excel templates config must be a list")

        base_dir = config_path.parent if config_path else self._project_root
        normalized: List[Dict[str, Any]] = []
        seen_titles = set()

        for idx, entry in enumerate(templates_data):
            if not isinstance(entry, dict):
                raise ValueError(f"Template entry at index {idx} must be an object")
            title = str(entry.get("title") or "").strip()
            desc = str(entry.get("desc") or "").strip()
            source_value = entry.get("source")

            if not title:
                raise ValueError(f"Template at index {idx} is missing 'title'")
            if not source_value:
                raise ValueError(f"Template '{title}' is missing 'source'")
            if title in seen_titles:
                raise ValueError(f"Duplicate template title: {title}")
            seen_titles.add(title)

            source_path = self._resolve_config_path(source_value, base_dir)
            if not source_path.exists():
                logger.warning("Excel template source not found: %s", self._friendly_path(source_path))

            normalized.append({"title": title, "desc": desc, "source": source_path})

        return normalized

    def _resolve_config_path(self, path_value: Any, base_dir: Optional[Path]) -> Path:
        """Resolve paths from config, supporting relative paths to the base directory."""
        candidate = Path(str(path_value))
        if str(candidate).startswith(("http://", "https://")):
            raise ValueError("Remote template sources are not supported; provide a local file path")
        if not candidate.is_absolute():
            candidate = (base_dir or self._project_root) / candidate
        return candidate.expanduser().resolve()

    def _friendly_path(self, path: Optional[Path]) -> str:
        """Return path relative to project root when possible."""
        if path is None:
            return ""
        try:
            return str(path.relative_to(self._project_root))
        except Exception:
            return str(path)

    def _ensure_templates_available(self) -> None:
        if self._template_error:
            raise ValueError(self._template_error)
        if not self.templates:
            hint = "No Excel templates configured"
            if self.templates_config_path:
                hint = f"No Excel templates configured (expected {self._friendly_path(self.templates_config_path)})"
            raise ValueError(hint)

    def _get_template_by_title(self, template_title: str) -> Dict[str, Any]:
        normalized = (template_title or "").strip()
        for tpl in self.templates:
            if tpl.get("title") == normalized:
                return tpl
        raise ValueError(f"Template not found: {template_title}")

    async def _resolve_template_target_path(
        self,
        directory: Optional[str],
        file_name: Optional[str],
        template_source: Path,
        template_title: str,
    ) -> Path:
        name_source = (file_name or template_title or "").strip()
        if not name_source:
            raise ValueError("file_name cannot be empty")

        logger.info(f"_resolve_template_target_path: directory={directory}, file_name={file_name}, template_title={template_title}")

        # Build virtual path string (not real Path object)
        if directory:
            # Validate directory first
            dir_path, allowed = await self.validator.validate_path(directory)
            if not allowed:
                raise ValueError(f"Path outside allowed directories: {directory}")
            # Convert to virtual path
            if self.validator.virtual_root:
                try:
                    dir_virtual = self.validator.real_to_virtual(dir_path)
                except ValueError:
                    # If conversion fails, use the directory as-is (might be already virtual)
                    dir_virtual = directory
            else:
                dir_virtual = str(dir_path)
            # Build virtual path string
            virtual_path_str = f"{dir_virtual.rstrip('/')}/{name_source}"
        else:
            # Use root virtual path
            virtual_path_str = f"/{name_source}"

        logger.info(f"_resolve_template_target_path: Constructed virtual_path_str={virtual_path_str}")

        # Add suffix if missing
        if not Path(virtual_path_str).suffix:
            suffix = template_source.suffix or ".xlsx"
            virtual_path_str = virtual_path_str + suffix
            logger.info(f"_resolve_template_target_path: Added suffix, new virtual_path_str={virtual_path_str}")

        # Validate and convert to real path
        target_path, allowed = await self.validator.validate_path(virtual_path_str)
        if not allowed:
            raise ValueError(f"Path outside allowed directories: {virtual_path_str}")

        self._ensure_supported_extension(target_path)
        logger.info(f"_resolve_template_target_path: Final target_path={target_path}")
        return target_path

    async def _resolve_target_directory(self, directory: Optional[str]) -> Path:
        if directory:
            dir_path, allowed = await self.validator.validate_path(directory)
            if not allowed:
                raise ValueError(f"Path outside allowed directories: {directory}")
        else:
            dir_path = self._default_directory()

        if not dir_path.exists():
            await anyio.to_thread.run_sync(lambda: dir_path.mkdir(parents=True, exist_ok=True))
        if not dir_path.is_dir():
            raise NotADirectoryError(f"Target directory is not a folder: {dir_path}")
        return dir_path

    def _default_directory(self) -> Path:
        if self.validator.virtual_root:
            return self.validator.virtual_root
        if self.validator.allowed_dirs:
            sorted_dirs = sorted(self.validator.allowed_dirs)
            if sorted_dirs:
                return Path(sorted_dirs[0])
        raise ValueError("No allowed directories available for template creation")

    async def _make_unique_path(self, path: Path) -> Path:
        """Return a non-conflicting path by appending (N) if needed.

        Note: This method expects path to already be a validated REAL path,
        not a virtual path that needs conversion.
        """
        # Path is already a real path from _resolve_template_target_path
        # Do NOT call validate_path again as it will treat it as virtual path
        # and cause path duplication!

        # Just verify it's within allowed directories
        abs_path = path.absolute()
        normalized = self.validator._normalize_case(str(abs_path))
        is_allowed = any(normalized.startswith(allowed_dir)
                        for allowed_dir in self.validator.allowed_dirs)

        if not is_allowed:
            raise ValueError(f"Path outside allowed directories: {path}")

        logger.info(f"_make_unique_path: input path={path}, abs_path={abs_path}")

        if not abs_path.exists():
            logger.info(f"_make_unique_path: Path doesn't exist, returning as-is")
            return abs_path

        # File exists, find a unique name
        stem = abs_path.stem
        suffix = abs_path.suffix
        parent = abs_path.parent
        counter = 1

        while True:
            candidate = parent / f"{stem}({counter}){suffix}"
            # Again, don't call validate_path, just check if it exists
            if not candidate.exists():
                logger.info(f"_make_unique_path: Found unique name: {candidate}")
                return candidate
            counter += 1
            if counter > 1000:  # Safety limit
                raise ValueError(f"Could not find unique path after 1000 attempts")

    async def _validated_excel_path(self, path: str, must_exist: bool = False) -> Path:
        """Validate Excel file path.

        Args:
            path: Virtual path string (NOT a Path object)
        """
        logger.debug(f"_validated_excel_path: Validating path: {path}, must_exist={must_exist}")
        
        if not isinstance(path, str):
            error_msg = (
                f"_validated_excel_path expects str (virtual path), got {type(path).__name__}. "
                "Do not pass Path objects - use virtual path strings like '/file.xlsx'"
            )
            logger.error(f"_validated_excel_path: {error_msg}")
            raise TypeError(error_msg)

        try:
            abs_path, allowed = await self.validator.validate_path(path)
            logger.debug(f"_validated_excel_path: Validation result - abs_path={abs_path}, allowed={allowed}")
            
            if not allowed:
                error_msg = f"Path outside allowed directories: {path}"
                logger.error(f"_validated_excel_path: {error_msg}")
                raise ValueError(error_msg)
            
            if must_exist and not abs_path.exists():
                error_msg = f"File not found: {path} (real path: {abs_path})"
                logger.error(f"_validated_excel_path: {error_msg}")
                raise FileNotFoundError(error_msg)
            
            logger.debug(f"_validated_excel_path: Checking file extension: {abs_path}")
            self._ensure_supported_extension(abs_path)
            logger.debug(f"_validated_excel_path: Path validation successful: {abs_path}")
            return abs_path
        except Exception as e:
            logger.error(f"_validated_excel_path: Validation failed for '{path}': {type(e).__name__}: {e}", exc_info=True)
            raise

    async def _load_workbook(
        self, path: Path, read_only: bool = True, data_only: bool = True
    ):
        return await anyio.to_thread.run_sync(
            partial(load_workbook, path, read_only=read_only, data_only=data_only)
        )

    def _get_sheet(self, wb, sheet: Optional[str]):
        """Get a worksheet from workbook.
        
        Args:
            wb: Workbook object
            sheet: Sheet name (None or empty string means use first sheet)
            
        Returns:
            Worksheet object
        """
        # Normalize sheet parameter: empty string -> None
        if sheet == "":
            sheet = None
        
        logger.debug(f"_get_sheet: Requested sheet: {sheet}, Available sheets: {wb.sheetnames}")
        
        if sheet:
            if sheet not in wb.sheetnames:
                available_sheets = ", ".join(wb.sheetnames) if wb.sheetnames else "(no sheets)"
                error_msg = (
                    f"Sheet '{sheet}' not found in workbook. "
                    f"Available sheets: {available_sheets}"
                )
                logger.error(f"_get_sheet: {error_msg}")
                raise ValueError(error_msg)
            logger.debug(f"_get_sheet: Returning sheet: {sheet}")
            return wb[sheet]
        
        # No sheet specified, use first sheet
        if not wb.sheetnames:
            error_msg = "Workbook has no sheets"
            logger.error(f"_get_sheet: {error_msg}")
            raise ValueError(error_msg)
        
        first_sheet_name = wb.sheetnames[0]
        logger.info(f"_get_sheet: No sheet specified, using first sheet: {first_sheet_name}")
        return wb[first_sheet_name]

    def _parse_range(self, ws, range_str: Optional[str]) -> Tuple[int, int, int, int]:
        """Parse a range string and return (min_row, max_row, min_col, max_col)."""
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        return self._parse_range_bounds(max_row, max_col, range_str)

    def _parse_range_bounds(
        self, max_row: int, max_col: int, range_str: Optional[str]
    ) -> Tuple[int, int, int, int]:
        """Parse a range string against provided bounds."""
        max_row = max_row or 1
        max_col = max_col or 1

        if not range_str:
            return 1, max_row, 1, max_col

        try:
            min_col, min_row, max_col_parsed, max_row_parsed = range_boundaries(range_str)
            return min_row, max_row_parsed, min_col, max_col_parsed
        except ValueError:
            pass

        # Handle column-only or row-only ranges
        if ":" in range_str:
            start, end = range_str.split(":", 1)
            if start.isdigit() and end.isdigit():
                min_row = int(start)
                max_row = int(end)
                return min_row, max_row, 1, max_col

            min_col = column_index_from_string(start)
            max_col = column_index_from_string(end)
            return 1, max_row, min_col, max_col

        # Single column without row numbers
        if range_str.isdigit():
            row_idx = int(range_str)
            return row_idx, row_idx, 1, max_col

        col_idx = column_index_from_string(range_str)
        return 1, max_row, col_idx, col_idx

    def _check_file_size(self, path: Path) -> None:
        """Validate file size against configured limit."""
        if not path.exists():
            return
        max_bytes = self.max_file_size_mb * 1024 * 1024
        size = path.stat().st_size
        if size > max_bytes:
            raise ValueError(
                f"File too large ({format_size(size)}), limit is {self.max_file_size_mb} MB"
            )

    def _ensure_supported_extension(self, path: Path) -> None:
        suffix = path.suffix.lower()
        if suffix and self.supported_formats and suffix not in self.supported_formats:
            raise ValueError(f"Unsupported file format: {suffix}")

    async def _resolve_output_path(
        self, source_path: Path, provided_output: Optional[str], default_suffix: str
    ) -> Path:
        """Resolve output path for conversion operations.

        Args:
            source_path: Source file path (already a validated REAL path)
            provided_output: User-provided output path (virtual path string) or None
            default_suffix: Default suffix to use if no output path provided

        Returns:
            Real path for output file
        """
        if provided_output:
            # User provided output path as virtual path string
            output_path, allowed = await self.validator.validate_path(provided_output)
            if not allowed:
                raise ValueError(f"Output path outside allowed directories: {provided_output}")
        else:
            # Generate default output path from source path
            # source_path is already a real path, so just change suffix
            output_path = source_path.with_suffix(default_suffix)

            # Verify the generated path is within allowed directories
            # Do NOT call validate_path as it would treat real path as virtual
            abs_path = output_path.absolute()
            normalized = self.validator._normalize_case(str(abs_path))
            is_allowed = any(normalized.startswith(allowed_dir)
                            for allowed_dir in self.validator.allowed_dirs)
            if not is_allowed:
                raise ValueError(f"Output path outside allowed directories: {output_path}")

            logger.info(f"_resolve_output_path: Generated default output_path={output_path}")

        if not output_path.parent.exists():
            await anyio.to_thread.run_sync(lambda: output_path.parent.mkdir(parents=True, exist_ok=True))

        return output_path

    def _apply_update(self, ws, update: Dict[str, Any]) -> int:
        """Apply a single update instruction to a worksheet."""
        # Support both "cell" and "range" parameter names for compatibility
        cell_ref = update.get("cell") or update.get("range")
        if not cell_ref:
            raise ValueError("Missing 'cell' or 'range' in update instruction")

        value = update.get("value")
        formula = update.get("formula")

        if ":" in cell_ref:
            min_row, max_row, min_col, max_col = self._parse_range(ws, cell_ref)
            rows_count = max_row - min_row + 1
            cols_count = max_col - min_col + 1
            values_matrix = self._expand_values(value, rows_count, cols_count, formula)

            for row_offset, row_values in enumerate(values_matrix):
                for col_offset, cell_value in enumerate(row_values):
                    target_row = min_row + row_offset
                    target_col = min_col + col_offset
                    cell_coord = f"{get_column_letter(target_col)}{target_row}"
                    ws[cell_coord] = cell_value
            return rows_count * cols_count

        if formula:
            ws[cell_ref] = formula
        else:
            ws[cell_ref] = value
        return 1

    def _expand_values(
        self,
        value: Any,
        rows: int,
        cols: int,
        formula: Optional[str],
    ) -> List[List[Any]]:
        """Expand the provided value into a matrix matching the target range."""
        if formula is not None:
            return [[formula for _ in range(cols)] for _ in range(rows)]

        if isinstance(value, list):
            # List of lists
            if value and isinstance(value[0], list):
                normalized: List[List[Any]] = []
                for row in value:
                    # Pad or trim to expected width
                    padded = list(row[:cols]) + [None] * max(0, cols - len(row))
                    normalized.append(padded[:cols])
                # Pad missing rows
                while len(normalized) < rows:
                    normalized.append([None] * cols)
                return normalized[:rows]

            # Flat list - fill row-major order
            flattened = list(value)
            normalized: List[List[Any]] = []
            idx = 0
            for _ in range(rows):
                row_values: List[Any] = []
                for _ in range(cols):
                    row_values.append(flattened[idx] if idx < len(flattened) else None)
                    idx += 1
                normalized.append(row_values)
            return normalized

        # Single scalar - broadcast
        return [[value for _ in range(cols)] for _ in range(rows)]

    def format_read_result(self, result: Dict[str, Any], output_format: str) -> str:
        """Public formatter wrapper for read_excel outputs."""
        fmt = normalize_output_format(output_format)
        result["format"] = fmt
        return format_read_result(result, fmt)

    # ---------- Advanced Excel operations inspired by excel-mcp-server ----------

    async def apply_formula(self, path: str, sheet: str, cell: str, formula: str) -> Dict[str, Any]:
        """Apply an Excel formula to a cell."""
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet}")
        # Basic syntax validation
        try:
            Tokenizer(formula)
        except Exception as e:
            raise ValueError(f"Invalid formula: {e}")
        ws = wb[sheet]
        ws[cell] = formula
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": f"Formula applied to {cell}", "path": self.validator.real_to_virtual(abs_path)}

    async def validate_formula_syntax(self, path: str, sheet: str, cell: str, formula: str) -> Dict[str, Any]:
        """Validate formula without applying."""
        await self._validated_excel_path(path, must_exist=True)
        try:
            Tokenizer(formula)
        except Exception as e:
            return {"success": False, "message": f"Invalid formula: {e}"}
        return {"success": True, "message": "Formula syntax looks valid"}

    async def format_range(
        self,
        path: str,
        sheet: str,
        start_cell: str,
        end_cell: Optional[str] = None,
        bold: bool = False,
        italic: bool = False,
        underline: bool = False,
        font_size: Optional[int] = None,
        font_color: Optional[str] = None,
        bg_color: Optional[str] = None,
        border_style: Optional[str] = None,
        border_color: Optional[str] = None,
        number_format: Optional[str] = None,
        alignment: Optional[str] = None,
        wrap_text: bool = False,
        merge_cells: bool = False,
        protection: Optional[Dict[str, Any]] = None,
        conditional_format: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        """Apply formatting to a cell range."""
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet}")
        ws = wb[sheet]
        rng = f"{start_cell}:{end_cell}" if end_cell else start_cell
        cells = ws[rng]

        font = Font(
            bold=bold or None,
            italic=italic or None,
            underline="single" if underline else None,
            size=font_size,
            color=font_color if font_color else None,
        )
        fill = PatternFill(fill_type="solid", fgColor=bg_color) if bg_color else None
        align_map = {
            "left": ("left", None),
            "center": ("center", None),
            "right": ("right", None),
            "top": (None, "top"),
            "middle": (None, "center"),
            "bottom": (None, "bottom"),
        }
        horizontal = vertical = None
        if alignment:
            horizontal, vertical = align_map.get(alignment, (None, None))
        align = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text or None)

        border = None
        if border_style:
            side = Side(style=border_style, color=border_color or "000000")
            border = Border(left=side, right=side, top=side, bottom=side)

        for row in cells:
            for c in row:
                c.font = font.copy() if font else c.font
                c.fill = fill.copy() if fill else c.fill
                c.alignment = align
                if border:
                    c.border = border
                if number_format:
                    c.number_format = number_format

        if merge_cells and end_cell:
            ws.merge_cells(rng)

        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": "Range formatted successfully", "path": self.validator.real_to_virtual(abs_path)}

    async def read_data_with_metadata(
        self,
        path: str,
        sheet: str,
        start_cell: str = "A1",
        end_cell: Optional[str] = None,
        preview_only: bool = False,
    ) -> Dict[str, Any]:
        """Read cells with metadata and validation info."""
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=True, data_only=False)
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet}")
        ws = wb[sheet]
        rng = f"{start_cell}:{end_cell}" if end_cell else start_cell
        min_col, min_row, max_col, max_row = range_boundaries(rng)
        cells_data = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for c in row:
                cells_data.append(
                    {
                        "address": c.coordinate,
                        "value": c.value,
                        "row": c.row,
                        "column": c.column,
                        "data_type": c.data_type,
                        "number_format": c.number_format,
                    }
                )
        return {
            "success": True,
            "sheet": sheet,
            "cells": cells_data if not preview_only else cells_data[: min(20, len(cells_data))],
            "path": self.validator.real_to_virtual(abs_path),
        }

    async def write_data_to_excel(
        self,
        path: str,
        sheet: str,
        data: List[List[Any]],
        start_cell: str = "A1",
    ) -> Dict[str, Any]:
        """Write a 2D list to sheet starting at start_cell without clearing existing data."""
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet}")
        ws = wb[sheet]
        start_col_idx, start_row_idx = range_boundaries(f"{start_cell}:{start_cell}")[:2]
        for r_idx, row in enumerate(data):
            for c_idx, value in enumerate(row):
                target = ws.cell(row=start_row_idx + r_idx, column=start_col_idx + c_idx)
                target.value = value
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": "Data written", "path": self.validator.real_to_virtual(abs_path)}

    async def create_workbook(self, path: str) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path)
        if abs_path.exists():
            raise FileExistsError(f"File already exists: {path}")
        wb = Workbook()
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": "Workbook created", "path": self.validator.real_to_virtual(abs_path)}

    async def create_worksheet(self, path: str, sheet_name: str) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        if sheet_name in wb.sheetnames:
            raise ValueError("Sheet already exists")
        wb.create_sheet(title=sheet_name)
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": f"Sheet '{sheet_name}' created", "path": self.validator.real_to_virtual(abs_path)}

    async def copy_worksheet(self, path: str, source_sheet: str, target_sheet: str) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        if source_sheet not in wb.sheetnames:
            raise ValueError("Source sheet not found")
        if target_sheet in wb.sheetnames:
            raise ValueError("Target sheet already exists")
        source = wb[source_sheet]
        new_ws = wb.copy_worksheet(source)
        new_ws.title = target_sheet
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": f"Sheet copied to '{target_sheet}'", "path": self.validator.real_to_virtual(abs_path)}

    async def delete_worksheet(self, path: str, sheet_name: str) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        if sheet_name not in wb.sheetnames:
            raise ValueError("Sheet not found")
        wb.remove(wb[sheet_name])
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": f"Sheet '{sheet_name}' deleted", "path": self.validator.real_to_virtual(abs_path)}

    async def rename_worksheet(self, path: str, old_name: str, new_name: str) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        if old_name not in wb.sheetnames:
            raise ValueError("Sheet not found")
        if new_name in wb.sheetnames:
            raise ValueError("Target name already exists")
        wb[old_name].title = new_name
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": f"Sheet renamed to '{new_name}'", "path": self.validator.real_to_virtual(abs_path)}

    async def get_workbook_metadata(self, path: str, include_ranges: bool = False) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=True, data_only=False)
        sheets = [{"name": s, "index": idx} for idx, s in enumerate(wb.sheetnames)]
        named_ranges = list(wb.defined_names.keys()) if include_ranges else []
        stat = abs_path.stat()
        return {
            "success": True,
            "path": self.validator.real_to_virtual(abs_path),
            "sheets": sheets,
            "named_ranges": named_ranges,
            "size": stat.st_size,
            "size_human": format_size(stat.st_size),
        }

    async def merge_cells(self, path: str, sheet: str, start_cell: str, end_cell: str) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        ws = wb[sheet]
        ws.merge_cells(f"{start_cell}:{end_cell}")
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": "Cells merged", "path": self.validator.real_to_virtual(abs_path)}

    async def unmerge_cells(self, path: str, sheet: str, start_cell: str, end_cell: str) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        ws = wb[sheet]
        ws.unmerge_cells(f"{start_cell}:{end_cell}")
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": "Cells unmerged", "path": self.validator.real_to_virtual(abs_path)}

    async def get_merged_cells(self, path: str, sheet: str) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=True, data_only=False)
        ws = wb[sheet]
        merged = [str(rng) for rng in ws.merged_cells.ranges]
        return {"success": True, "ranges": merged, "path": self.validator.real_to_virtual(abs_path)}

    async def copy_range(
        self,
        path: str,
        sheet: str,
        source_start: str,
        source_end: str,
        target_start: str,
        target_sheet: Optional[str] = None,
    ) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        src_ws = wb[sheet]
        tgt_ws = wb[target_sheet] if target_sheet else src_ws
        src_min_col, src_min_row, src_max_col, src_max_row = range_boundaries(f"{source_start}:{source_end}")
        tgt_start_col, tgt_start_row = range_boundaries(f"{target_start}:{target_start}")[:2]
        for r_off, row in enumerate(src_ws.iter_rows(min_row=src_min_row, max_row=src_max_row, min_col=src_min_col, max_col=src_max_col)):
            for c_off, cell in enumerate(row):
                tgt = tgt_ws.cell(row=tgt_start_row + r_off, column=tgt_start_col + c_off)
                tgt.value = cell.value
                tgt._style = cell._style
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": "Range copied", "path": self.validator.real_to_virtual(abs_path)}

    async def delete_range(
        self,
        path: str,
        sheet: str,
        start_cell: str,
        end_cell: str,
        shift_direction: str = "up",
    ) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        ws = wb[sheet]
        min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")
        if shift_direction in ("up", "down"):
            ws.delete_rows(min_row, max_row - min_row + 1)
        else:
            ws.delete_cols(min_col, max_col - min_col + 1)
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": "Range deleted", "path": self.validator.real_to_virtual(abs_path)}

    async def validate_excel_range(self, path: str, sheet: str, start_cell: str, end_cell: Optional[str] = None) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=True, data_only=False)
        if sheet not in wb.sheetnames:
            raise ValueError("Sheet not found")
        rng = start_cell if not end_cell else f"{start_cell}:{end_cell}"
        try:
            range_boundaries(rng)
        except ValueError as e:
            return {"success": False, "message": str(e)}
        return {"success": True, "message": "Range is valid"}

    async def get_data_validation_info(self, path: str, sheet: str) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=True, data_only=False)
        ws = wb[sheet]
        validations = []
        for dv in ws.data_validations.dataValidation:
            validations.append(
                {
                    "type": dv.type,
                    "sqref": str(dv.sqref),
                    "formula1": dv.formula1,
                    "formula2": dv.formula2,
                    "allow_blank": dv.allow_blank,
                    "showErrorMessage": dv.showErrorMessage,
                }
            )
        return {"success": True, "validations": validations, "path": self.validator.real_to_virtual(abs_path)}

    async def insert_rows(self, path: str, sheet: str, start_row: int, count: int = 1) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        ws = wb[sheet]
        ws.insert_rows(start_row, count)
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": "Rows inserted", "path": self.validator.real_to_virtual(abs_path)}

    async def insert_columns(self, path: str, sheet: str, start_col: int, count: int = 1) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        ws = wb[sheet]
        ws.insert_cols(start_col, count)
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": "Columns inserted", "path": self.validator.real_to_virtual(abs_path)}

    async def delete_rows(self, path: str, sheet: str, start_row: int, count: int = 1) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        ws = wb[sheet]
        ws.delete_rows(start_row, count)
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": "Rows deleted", "path": self.validator.real_to_virtual(abs_path)}

    async def delete_columns(self, path: str, sheet: str, start_col: int, count: int = 1) -> Dict[str, Any]:
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        ws = wb[sheet]
        ws.delete_cols(start_col, count)
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": "Columns deleted", "path": self.validator.real_to_virtual(abs_path)}

    async def create_chart(
        self,
        path: str,
        sheet: str,
        data_range: str,
        chart_type: str,
        target_cell: str,
        title: str = "",
        x_axis: str = "",
        y_axis: str = "",
    ) -> Dict[str, Any]:
        """Create a basic chart from a data range."""
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        ws = wb[sheet]
        min_col, min_row, max_col, max_row = range_boundaries(data_range)
        data_ref = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
        chart_type_lower = chart_type.lower()
        if chart_type_lower in ("column", "bar"):
            chart = BarChart()
        elif chart_type_lower == "line":
            chart = LineChart()
        elif chart_type_lower == "pie":
            chart = PieChart()
        elif chart_type_lower == "scatter":
            chart = ScatterChart()
            chart.series = [Series(data_ref, title=title or None)]
        else:
            raise ValueError("Unsupported chart type")
        if chart_type_lower != "scatter":
            chart.add_data(data_ref, titles_from_data=True)
        chart.title = title
        chart.x_axis.title = x_axis
        chart.y_axis.title = y_axis
        ws.add_chart(chart, target_cell)
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": "Chart created", "path": self.validator.real_to_virtual(abs_path)}

    async def create_table(
        self,
        path: str,
        sheet: str,
        data_range: str,
        table_name: Optional[str] = None,
        table_style: str = "TableStyleMedium9",
    ) -> Dict[str, Any]:
        from openpyxl.worksheet.table import Table, TableStyleInfo

        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=False)
        ws = wb[sheet]
        tbl = Table(displayName=table_name or f"Table_{len(ws._tables)+1}", ref=data_range)
        style = TableStyleInfo(name=table_style, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tbl.tableStyleInfo = style
        ws.add_table(tbl)
        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": "Table created", "path": self.validator.real_to_virtual(abs_path)}

    async def create_pivot_table(
        self,
        path: str,
        sheet: str,
        data_range: str,
        rows: List[str],
        values: List[str],
        columns: Optional[List[str]] = None,
        agg_func: str = "mean",
        target_sheet: Optional[str] = None,
        target_cell: str = "A1",
    ) -> Dict[str, Any]:
        """Simplified pivot: load data into memory and aggregate."""
        abs_path = await self._validated_excel_path(path, must_exist=True)
        wb = await self._load_workbook(abs_path, read_only=False, data_only=True)
        src_ws = wb[sheet]
        min_col, min_row, max_col, max_row = range_boundaries(data_range)
        headers = [cell.value for cell in src_ws[min_row]]
        data_rows = []
        for row in src_ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            data_rows.append(dict(zip(headers, row)))

        def agg(values_list):
            vals = [v for v in values_list if isinstance(v, (int, float))]
            if not vals:
                return None
            if agg_func == "sum":
                return sum(vals)
            if agg_func == "max":
                return max(vals)
            if agg_func == "min":
                return min(vals)
            if agg_func == "count":
                return len(vals)
            # default mean
            return sum(vals) / len(vals)

        pivot = {}
        for row_dict in data_rows:
            row_key = tuple(row_dict.get(r) for r in rows)
            col_key = tuple(row_dict.get(c) for c in (columns or [])) if columns else ()
            pivot.setdefault(row_key, {}).setdefault(col_key, {val: [] for val in values})
            for val in values:
                pivot[row_key][col_key][val].append(row_dict.get(val))

        out_rows = []
        for row_key, col_map in pivot.items():
            for col_key, val_map in col_map.items():
                entry = {r: rv for r, rv in zip(rows, row_key)}
                if columns:
                    for c, cv in zip(columns, col_key):
                        entry[c] = cv
                for val in values:
                    entry[val] = agg(val_map[val])
                out_rows.append(entry)

        tgt_ws = wb[target_sheet] if target_sheet and target_sheet in wb.sheetnames else wb.create_sheet(target_sheet or "Pivot")
        start_col_idx, start_row_idx = range_boundaries(f"{target_cell}:{target_cell}")[:2]
        # headers
        headers_out = list(rows) + (columns or []) + values
        for c_idx, h in enumerate(headers_out):
            tgt_ws.cell(row=start_row_idx, column=start_col_idx + c_idx, value=h)
        for r_offset, row in enumerate(out_rows, start=1):
            for c_idx, h in enumerate(headers_out):
                tgt_ws.cell(row=start_row_idx + r_offset, column=start_col_idx + c_idx, value=row.get(h))

        await anyio.to_thread.run_sync(wb.save, abs_path)
        return {"success": True, "message": "Pivot table created", "path": self.validator.real_to_virtual(abs_path), "sheet": tgt_ws.title}

    def _is_csv(self, path: Path) -> bool:
        return path.suffix.lower() == ".csv"

    # ---------- CSV helpers ----------

    async def _read_csv_file(
        self,
        path: Path,
        sheet: Optional[str],
        range_str: Optional[str],
        max_rows: Optional[int],
        output_format: str,
    ) -> Dict[str, Any]:
        rows = await read_csv(path)
        total_rows_available = len(rows)
        total_cols_available = max((len(r) for r in rows), default=0)

        min_row, max_row, min_col, max_col = self._parse_range_bounds(
            total_rows_available or 1, total_cols_available or 1, range_str
        )

        sliced_rows = self._slice_rows(rows, min_row, max_row, min_col, max_col)

        row_limit = max_rows or self.default_max_rows
        row_limit = row_limit if row_limit > 0 else self.default_max_rows
        truncated = len(sliced_rows) > row_limit
        if truncated:
            sliced_rows = sliced_rows[:row_limit]

        available_rows = max(0, min(max_row, total_rows_available) - min_row + 1)
        sheet_name = sheet or path.stem or "Sheet1"
        result: Dict[str, Any] = {
            "sheet": sheet_name,
            "sheets": [sheet_name],  # CSV 只有一个 sheet
            "rows": sliced_rows,
            "total_rows": available_rows,
            "returned_rows": len(sliced_rows),
            "truncated": truncated,
        }
        return result

    def _slice_rows(
        self,
        rows: List[List[Any]],
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int,
    ) -> List[List[Any]]:
        """Slice raw CSV rows using Excel-like coordinates."""
        sliced: List[List[Any]] = []
        for idx, row in enumerate(rows, start=1):
            if idx < min_row or idx > max_row:
                continue
            padded = list(row)
            if len(padded) < max_col:
                padded.extend([None] * (max_col - len(padded)))
            sliced.append(padded[min_col - 1 : max_col])
        return sliced
